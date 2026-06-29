"""
Excel Updater for Roy G's Liquor Inventory
- Reads On Hand quantities from Inventory Locations sheet
- Writes Provi Price, Twin Price, Best Price, and Best Vendor columns
- Highlights in green whichever vendor has the lower price
"""

import logging
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

log = logging.getLogger(__name__)

MASTER_SHEET    = "Master Order Guide"
INVENTORY_SHEET = "Inventory Locations"

# Existing column indices (1-based)
LEFT_ITEM_COL    = 1   # A
LEFT_ONHAND_COL  = 5   # E
LEFT_PRICE_COL   = 7   # G  Cost Per Bottle
LEFT_VENDOR_COL  = 9   # I

RIGHT_ITEM_COL   = 10  # J
RIGHT_ONHAND_COL = 14  # N
RIGHT_PRICE_COL  = 16  # P
RIGHT_VENDOR_COL = 18  # R

DATA_START_ROW = 5

# New comparison columns (appended after existing columns)
# Left block comparison: cols S(19), T(20), U(21), V(22)
LEFT_PROVI_COL      = 20  # T
LEFT_TWIN_COL       = 21  # U
LEFT_BEST_PRICE_COL = 22  # V
LEFT_BEST_VEN_COL   = 23  # W

# Right block comparison: cols X(24), Y(25), Z(26), AA(27)
RIGHT_PROVI_COL      = 24  # X
RIGHT_TWIN_COL       = 25  # Y
RIGHT_BEST_PRICE_COL = 26  # Z
RIGHT_BEST_VEN_COL   = 27  # AA

GREEN_FILL  = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL    = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
CLEAR_FILL  = PatternFill(fill_type=None)
BOLD        = Font(bold=True)


def _normalize(name):
    if not name:
        return ""
    return str(name).lower().strip()


def build_inventory_map(wb):
    ws = wb[INVENTORY_SHEET]
    inventory = {}
    group_starts = [1, 4, 7, 10, 13]
    skip = {
        "column 1 mid", "column 2 mid", "column 3 mid",
        "column 4 mid", "column 5 mid",
        "column 1 low", "column 2 low", "column 3 low",
        "column 4 low", "column 5 low",
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        for start in group_starts:
            idx = start - 1
            if idx >= len(row):
                continue
            name = row[idx]
            qty_idx = idx + 2
            qty = row[qty_idx] if qty_idx < len(row) else None
            if name and _normalize(name) not in skip:
                try:
                    quantity = float(qty) if qty is not None else 0
                except (ValueError, TypeError):
                    quantity = 0
                inventory[_normalize(name)] = quantity
    log.info(f"Loaded {len(inventory)} items from Inventory Locations.")
    return inventory


def find_best_match(item_name, price_dict):
    norm = _normalize(item_name)
    if norm in price_dict:
        return price_dict[norm]
    for key, price in price_dict.items():
        if key in norm or norm in key:
            return price
    return None


def write_comparison_headers(ws):
    """Write headers for the new comparison columns."""
    header_row = 3

    # Left block headers
    ws.cell(header_row, LEFT_PROVI_COL,      "Provi Price").font  = BOLD
    ws.cell(header_row, LEFT_TWIN_COL,       "Twin Price").font   = BOLD
    ws.cell(header_row, LEFT_BEST_PRICE_COL, "Best Price").font   = BOLD
    ws.cell(header_row, LEFT_BEST_VEN_COL,   "Best Vendor").font  = BOLD

    # Right block headers
    ws.cell(header_row, RIGHT_PROVI_COL,      "Provi Price").font  = BOLD
    ws.cell(header_row, RIGHT_TWIN_COL,       "Twin Price").font   = BOLD
    ws.cell(header_row, RIGHT_BEST_PRICE_COL, "Best Price").font   = BOLD
    ws.cell(header_row, RIGHT_BEST_VEN_COL,   "Best Vendor").font  = BOLD


def apply_comparison(ws, row_idx, item_name, provi_price, twin_price,
                     provi_col, twin_col, best_price_col, best_ven_col,
                     onhand_col, inventory):
    """Write price comparison data and highlight best vendor."""

    # Write individual prices
    p_cell = ws.cell(row_idx, provi_col)
    t_cell = ws.cell(row_idx, twin_col)
    b_cell = ws.cell(row_idx, best_price_col)
    v_cell = ws.cell(row_idx, best_ven_col)

    p_cell.value = provi_price
    t_cell.value = twin_price

    # Determine best price
    if provi_price and twin_price:
        if provi_price <= twin_price:
            b_cell.value = provi_price
            v_cell.value = "Provi"
            p_cell.fill = GREEN_FILL
            t_cell.fill = RED_FILL
        else:
            b_cell.value = twin_price
            v_cell.value = "Twin"
            t_cell.fill = GREEN_FILL
            p_cell.fill = RED_FILL
        b_cell.fill = GREEN_FILL
        v_cell.font = BOLD

        # Also update the main Cost Per Bottle with the best price
        main_price_col = LEFT_PRICE_COL if provi_col == LEFT_PROVI_COL else RIGHT_PRICE_COL
        ws.cell(row_idx, main_price_col).value = b_cell.value

    elif provi_price:
        b_cell.value = provi_price
        v_cell.value = "Provi"
        p_cell.fill = YELLOW_FILL
    elif twin_price:
        b_cell.value = twin_price
        v_cell.value = "Twin"
        t_cell.fill = YELLOW_FILL

    # Update On Hand from inventory
    qty = inventory.get(_normalize(item_name))
    if qty is not None:
        ws.cell(row_idx, onhand_col).value = qty


def update_excel(filepath, combined_prices, output_path=None):
    """
    combined_prices: {product_name: {"provi": float|None, "twin": float|None}}
    """
    wb = load_workbook(filepath)
    ws = wb[MASTER_SHEET]

    inventory = build_inventory_map(wb)

    # Normalize price keys
    norm_prices = {_normalize(k): v for k, v in combined_prices.items()}

    write_comparison_headers(ws)

    updated = 0
    skipped = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        # LEFT BLOCK
        left_name = ws.cell(row_idx, LEFT_ITEM_COL).value
        if left_name and str(left_name).strip():
            match = find_best_match(left_name, norm_prices)
            if match:
                apply_comparison(
                    ws, row_idx, left_name,
                    match.get("provi"), match.get("twin"),
                    LEFT_PROVI_COL, LEFT_TWIN_COL,
                    LEFT_BEST_PRICE_COL, LEFT_BEST_VEN_COL,
                    LEFT_ONHAND_COL, inventory
                )
                updated += 1
            else:
                skipped += 1

        # RIGHT BLOCK
        right_name = ws.cell(row_idx, RIGHT_ITEM_COL).value
        if right_name and str(right_name).strip():
            match = find_best_match(right_name, norm_prices)
            if match:
                apply_comparison(
                    ws, row_idx, right_name,
                    match.get("provi"), match.get("twin"),
                    RIGHT_PROVI_COL, RIGHT_TWIN_COL,
                    RIGHT_BEST_PRICE_COL, RIGHT_BEST_VEN_COL,
                    RIGHT_ONHAND_COL, inventory
                )
                updated += 1
            else:
                skipped += 1

    # Update date
    ws["D1"] = str(date.today())

    save_path = output_path or filepath
    wb.save(save_path)
    log.info(f"Saved to {save_path} | Updated: {updated} | Skipped: {skipped}")

    return {"updated": updated, "skipped": skipped}
