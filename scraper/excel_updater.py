"""
Excel Updater for Roy G's Liquor Inventory
- Reads On Hand quantities from "Inventory Locations" sheet
- Writes updated prices into "Master Order Guide" from Provi / Twin B2B
"""

import logging
from datetime import date
from openpyxl import load_workbook

log = logging.getLogger(__name__)

MASTER_SHEET    = "Master Order Guide"
INVENTORY_SHEET = "Inventory Locations"

# Column indices in Master Order Guide (1-based)
# Layout: two side-by-side blocks sharing the same header row (row 3)
# Left block:  A=item, B=level, C=size, D=par, E=on_hand, F=need_to_order, G=cost_per_bottle, H=extended_value, I=vendor
# Right block: J=item, K=level, L=size, M=par, N=on_hand, O=need_to_order, P=cost_per_bottle, Q=extended_value, R=vendor

LEFT_ITEM_COL   = 1   # A
LEFT_ONHAND_COL = 5   # E
LEFT_PRICE_COL  = 7   # G  Cost Per Bottle
LEFT_VENDOR_COL = 9   # I

RIGHT_ITEM_COL   = 10  # J
RIGHT_ONHAND_COL = 14  # N
RIGHT_PRICE_COL  = 16  # P
RIGHT_VENDOR_COL = 18  # R

DATA_START_ROW = 5     # Row 5 is first data row (rows 1-4 are header/title)


def _normalize(name):
    """Lowercase + strip for fuzzy matching."""
    if not name:
        return ""
    return str(name).lower().strip()


def build_inventory_map(wb):
    """
    Read Inventory Locations sheet.
    Returns dict: {normalized_product_name: quantity}
    Currently the sheet only has names; quantity column (col B, E, H, K, N per group)
    is expected to be filled in by the user adjacent to each item name.
    Col layout per 3-col group: [item_name, blank/label, quantity]
    Groups start at cols: A(1), D(4), G(7), J(10), M(13)
    """
    ws = wb[INVENTORY_SHEET]
    inventory = {}

    # Each "column group" starts at these 1-based column indices
    group_starts = [1, 4, 7, 10, 13]

    for row in ws.iter_rows(min_row=2, values_only=True):
        for start in group_starts:
            idx = start - 1  # 0-based index into row tuple
            if idx >= len(row):
                continue
            name = row[idx]
            # Quantity is expected in the 3rd column of each group (idx+2)
            qty_idx = idx + 2
            qty = row[qty_idx] if qty_idx < len(row) else None

            if name and str(name).strip() and str(name).strip() not in (
                "Column 1 Mid", "Column 2 Mid", "Column 3 Mid",
                "Column 4 Mid", "Column 5 Mid",
                "Column 1 Low", "Column 2 Low", "Column 3 Low",
                "Column 4 Low", "Column 5 Low",
            ):
                try:
                    quantity = float(qty) if qty is not None else 0
                except (ValueError, TypeError):
                    quantity = 0
                inventory[_normalize(name)] = quantity

    log.info(f"Loaded {len(inventory)} items from Inventory Locations.")
    return inventory


def find_best_match(item_name, prices_dict):
    """
    Try to match an Excel item name to a price result key.
    Exact normalized match first, then partial/substring match.
    """
    norm = _normalize(item_name)
    if norm in prices_dict:
        return prices_dict[norm]

    # Try substring match both ways
    for key, price in prices_dict.items():
        if key in norm or norm in key:
            return price

    return None


def update_excel(filepath, provi_prices, twin_prices, output_path=None):
    """
    Load the workbook, update prices and on-hand quantities, save.

    provi_prices / twin_prices: {product_name_str: float_or_None}
    """
    wb = load_workbook(filepath)
    ws = wb[MASTER_SHEET]

    # Normalize price dicts for matching
    provi_norm = {_normalize(k): v for k, v in provi_prices.items() if v is not None}
    twin_norm  = {_normalize(k): v for k, v in twin_prices.items()  if v is not None}

    inventory = build_inventory_map(wb)

    updated_prices  = 0
    updated_onhand  = 0
    skipped         = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        # ---- LEFT BLOCK ----
        left_item_cell   = row[LEFT_ITEM_COL - 1]
        left_price_cell  = row[LEFT_PRICE_COL - 1]
        left_onhand_cell = row[LEFT_ONHAND_COL - 1]
        left_vendor_cell = row[LEFT_VENDOR_COL - 1]

        left_name   = left_item_cell.value
        left_vendor = str(left_vendor_cell.value or "").strip().lower()

        if left_name and str(left_name).strip():
            # Update price based on vendor
            if "twin" in left_vendor:
                price = find_best_match(left_name, twin_norm)
            else:
                price = find_best_match(left_name, provi_norm)

            if price is not None:
                left_price_cell.value = price
                updated_prices += 1
            else:
                skipped += 1

            # Update On Hand from inventory
            qty = inventory.get(_normalize(left_name))
            if qty is not None:
                left_onhand_cell.value = qty
                updated_onhand += 1

        # ---- RIGHT BLOCK ----
        right_item_cell   = row[RIGHT_ITEM_COL - 1]
        right_price_cell  = row[RIGHT_PRICE_COL - 1]
        right_onhand_cell = row[RIGHT_ONHAND_COL - 1]
        right_vendor_cell = row[RIGHT_VENDOR_COL - 1]

        right_name   = right_item_cell.value
        right_vendor = str(right_vendor_cell.value or "").strip().lower()

        if right_name and str(right_name).strip():
            if "twin" in right_vendor:
                price = find_best_match(right_name, twin_norm)
            else:
                price = find_best_match(right_name, provi_norm)

            if price is not None:
                right_price_cell.value = price
                updated_prices += 1
            else:
                skipped += 1

            qty = inventory.get(_normalize(right_name))
            if qty is not None:
                right_onhand_cell.value = qty
                updated_onhand += 1

    # Update the date cell (row 1, col C = "Date:")
    ws["D1"] = str(date.today())

    save_path = output_path or filepath
    wb.save(save_path)
    log.info(f"Saved to {save_path}")
    log.info(f"  Prices updated : {updated_prices}")
    log.info(f"  On Hand updated: {updated_onhand}")
    log.info(f"  Skipped (no match): {skipped}")

    return {
        "prices_updated": updated_prices,
        "onhand_updated": updated_onhand,
        "skipped": skipped,
    }
