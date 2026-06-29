"""
Main entry point for the Roy G's Spirits Price Tracker.

Usage:
    python main.py --file path/to/spreadsheet.xlsx [--output path/to/output.xlsx]

The script:
1. Reads the Excel file to extract which products go to which vendor
2. Scrapes current prices from Provi (provi.com) and Twin B2B (twinb2b.com)
3. Updates the Excel file with new prices and on-hand quantities
4. Saves the updated file (overwrites original unless --output is specified)
"""

import argparse
import logging
import sys
from openpyxl import load_workbook

from scraper import scrape_provi, scrape_twin
from excel_updater import update_excel, DATA_START_ROW, MASTER_SHEET
from excel_updater import (
    LEFT_ITEM_COL, LEFT_VENDOR_COL,
    RIGHT_ITEM_COL, RIGHT_VENDOR_COL,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def extract_products_by_vendor(filepath):
    """
    Read the Master Order Guide and split items into provi_products / twin_products lists.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb[MASTER_SHEET]

    provi_products = []
    twin_products  = []

    skip_values = {
        "vodka", "rum", "whiskey", "gin", "wine", "cordials",
        "beer/seltzer/rtd", "n/a bev", "garnish", "agave/tequila",
        "liquor\nlevel", "liquor level", "size", "par", "on hand",
        "need to order", "cost per bottle", "extended value", "vendor",
        "roy g's liquor inventory", "date:",
    }

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        # Left block
        left_name   = row[LEFT_ITEM_COL - 1]
        left_vendor = str(row[LEFT_VENDOR_COL - 1] or "").strip().lower()
        if left_name and str(left_name).strip().lower() not in skip_values:
            name = str(left_name).strip()
            if "twin" in left_vendor:
                twin_products.append(name)
            elif left_vendor:  # specs / provi / southern etc
                provi_products.append(name)

        # Right block
        right_name   = row[RIGHT_ITEM_COL - 1]
        right_vendor = str(row[RIGHT_VENDOR_COL - 1] or "").strip().lower()
        if right_name and str(right_name).strip().lower() not in skip_values:
            name = str(right_name).strip()
            if "twin" in right_vendor:
                twin_products.append(name)
            elif right_vendor:
                provi_products.append(name)

    # Deduplicate while preserving order
    provi_products = list(dict.fromkeys(provi_products))
    twin_products  = list(dict.fromkeys(twin_products))

    log.info(f"Products to fetch from Provi : {len(provi_products)}")
    log.info(f"Products to fetch from Twin  : {len(twin_products)}")

    return provi_products, twin_products


def main():
    parser = argparse.ArgumentParser(description="Roy G's Spirits Price Tracker")
    parser.add_argument("--file",   required=True, help="Path to the Excel spreadsheet")
    parser.add_argument("--output", default=None,  help="Output path (defaults to overwrite input)")
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("Roy G's Spirits Price Tracker — Starting")
    log.info("=" * 60)

    # Step 1: Read product lists from the spreadsheet
    provi_products, twin_products = extract_products_by_vendor(args.file)

    # Step 2: Scrape prices
    log.info("Scraping Provi prices...")
    provi_prices = scrape_provi(provi_products)

    log.info("Scraping Twin B2B prices...")
    twin_prices = scrape_twin(twin_products)

    # Step 3: Update Excel
    log.info("Updating Excel spreadsheet...")
    stats = update_excel(
        filepath=args.file,
        provi_prices=provi_prices,
        twin_prices=twin_prices,
        output_path=args.output,
    )

    log.info("=" * 60)
    log.info("Done!")
    log.info(f"  Prices updated : {stats['prices_updated']}")
    log.info(f"  On Hand synced : {stats['onhand_updated']}")
    log.info(f"  Not matched    : {stats['skipped']}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
